#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_competitor_analysis():
    """Rakip Analizi Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Rakip Analizi"

    # Başlıklar
    headers = ["Özellik", "QuadroAIPilot", "ChatGPT", "Gemini", "NotebookLM",
               "Claude.ai", "Perplexity", "Copilot", "Character.AI", "Açıklama"]

    # Veri satırları
    data = [
        ["Sesli Komut", "✅", "✅", "✅", "❌", "❌", "✅", "✅", "✅", "Mikrofonla komut verme"],
        ["Karşılıklı Sesli Sohbet", "❌", "✅", "✅", "❌", "❌", "✅", "✅", "✅", "Telefon görüşmesi gibi konuşma"],
        ["Sohbet Geçmişi Kaydetme", "❌", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "Dünkü konuşmaları okuma"],
        ["Sohbet Export (PDF/Word)", "❌", "✅", "❌", "✅", "❌", "❌", "✅", "❌", "Konuşmaları dosyaya kaydetme"],
        ["Dosya Yükleme (PDF/Word)", "❌", "✅", "✅", "✅", "✅", "✅", "✅", "❌", "Belge yükleyip analiz ettirme"],
        ["Web/İnternet Araştırması", "⚠️", "✅", "✅", "✅", "❌", "✅", "✅", "❌", "Güncel internet bilgisi"],
        ["Kaynak Gösterme", "❌", "⚠️", "❌", "✅", "❌", "✅", "❌", "❌", "Bilginin nereden geldiğini gösterme"],
        ["Mobil Uygulama", "❌", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "iOS/Android uygulaması"],
        ["Kamera/Ekran Paylaşımı", "❌", "⚠️", "✅", "❌", "❌", "⚠️", "✅", "❌", "Ekranı/kamerayı gösterip soru sorma"],
        ["Çoklu Dil (İngilizce)", "⚠️", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "İngilizce ve diğer diller"],
        ["Projeler/Klasörler", "❌", "✅", "⚠️", "✅", "✅", "❌", "⚠️", "❌", "Konuşmaları klasörlere ayırma"],
        ["Plugin/Eklenti Sistemi", "❌", "✅", "✅", "✅", "✅", "❌", "✅", "❌", "Yeni özellikler ekleme"],
        ["Windows Entegrasyonu", "✅", "❌", "⚠️", "❌", "❌", "❌", "✅", "❌", "Windows komutları çalıştırma"],
        ["Outlook/Mail Entegrasyonu", "✅", "❌", "⚠️", "❌", "❌", "❌", "✅", "❌", "Mail okuma/gönderme"],
        ["Türkçe Dikte", "✅", "⚠️", "⚠️", "❌", "❌", "⚠️", "⚠️", "⚠️", "Türkçe sesli yazı yazdırma"],
        ["Komut Modu", "✅", "❌", "❌", "❌", "❌", "❌", "⚠️", "❌", "Sistem komutları (kopyala, aç, kapat)"],
        ["Yazı Modu", "✅", "❌", "❌", "❌", "❌", "❌", "❌", "❌", "Sesli dikte ile yazı yazma"],
        ["Context Window", "⚠️", "8K-400K", "1M-2M", "200K", "200K", "-", "-", "-", "Kaç kelime hatırlayabiliyor"],
        ["Streaming Yanıt", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "✅", "Cevabı parça parça gösterme"],
        ["Ücretsiz Plan", "✅", "✅", "✅", "✅", "⚠️", "✅", "✅", "✅", "Ücretsiz kullanılabiliyor mu"],
    ]

    # Başlık satırını yaz
    ws.append(headers)

    # Veri satırlarını yaz
    for row in data:
        ws.append(row)

    # Stil tanımlamaları
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Başlık satırı formatla
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Veri satırlarını formatla
    for row_idx in range(2, len(data) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Değere göre renklendirme
            if cell.value == "✅":
                cell.fill = green_fill
                cell.font = Font(size=14)
            elif cell.value == "❌":
                cell.fill = red_fill
                cell.font = Font(size=14)
            elif cell.value == "⚠️":
                cell.fill = orange_fill
                cell.font = Font(size=14)

            # İlk sütun (özellik adı) sola hizalı ve kalın
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)

            # Son sütun (açıklama) sola hizalı
            if col_idx == len(headers):
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Sütun genişliklerini ayarla
    column_widths = [30, 15, 12, 12, 15, 12, 12, 12, 15, 35]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Satır yüksekliğini ayarla
    ws.row_dimensions[1].height = 30
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 25

    # Açıklama satırları ekle
    last_row = len(data) + 2
    ws.append([])
    ws.append(["Sembol Açıklaması:"])
    ws.cell(row=last_row + 2, column=1).font = Font(bold=True, size=11)

    ws.append(["✅ = Tam destekli / Çok iyi çalışıyor"])
    ws.append(["⚠️ = Kısmi destek / Sınırlı çalışıyor"])
    ws.append(["❌ = Yok / Desteklenmiyor"])
    ws.append(["- = Bilgi bulunamadı"])

    # Kaydet
    wb.save("/mnt/c/Users/serkan/source/repos/QuadroAIPilot setup so so outlook not setup deneme2/QuadroAIPilot_Rakip_Analizi.xlsx")
    print("✅ QuadroAIPilot_Rakip_Analizi.xlsx oluşturuldu")

def create_priority_analysis():
    """Eksikler Öncelik Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Eksikler Öncelik"

    # Başlıklar
    headers = ["Özellik", "Öncelik", "Süre (Hafta)", "Zorluk", "Kullanıcı Etkisi",
               "Rakiplerde Var mı", "Açıklama"]

    # Veri satırları
    data = [
        ["Sohbet Geçmişi Kaydetme", "🔴 Kritik", "2", "Orta", "Çok Yüksek", "7/7", "Kullanıcılar dünkü konuşmaları okuyamıyor - temel beklenti"],
        ["Karşılıklı Sesli Sohbet", "🔴 Kritik", "4", "Yüksek", "Çok Yüksek", "5/7", "Telefon görüşmesi gibi konuşma - modern AI'larda standart"],
        ["Web/İnternet Araştırması", "🔴 Kritik", "3", "Orta", "Çok Yüksek", "6/7", "Güncel bilgi alma - şu anda kısmi çalışıyor, tam entegrasyon gerekli"],
        ["Kaynak Gösterme", "🔴 Kritik", "1", "Düşük", "Yüksek", "2/7", "Bilginin nereden geldiğini gösterme - güvenilirlik için kritik"],
        ["Projeler/Klasörler", "🟡 Önemli", "2", "Orta", "Yüksek", "4/7", "Konuşmaları organize etme - iş kullanıcıları için önemli"],
        ["Dosya Yükleme (PDF/Word)", "🟡 Önemli", "3", "Orta", "Yüksek", "6/7", "Belge analizi - profesyonel kullanım için gerekli"],
        ["Sohbet Export (PDF/Word)", "🟡 Önemli", "1", "Düşük", "Orta", "3/7", "Konuşmaları kaydetme - raporlama için yararlı"],
        ["Kamera/Ekran Paylaşımı", "🟡 Önemli", "5", "Yüksek", "Yüksek", "3/7", "Görsel analiz - destek ve eğitim için önemli"],
        ["Plugin/Eklenti Sistemi", "🟡 Önemli", "6", "Çok Yüksek", "Orta", "5/7", "Genişletilebilirlik - uzun vadeli büyüme için önemli"],
        ["Çoklu Dil (İngilizce)", "🟢 İsteğe Bağlı", "2", "Düşük", "Orta", "7/7", "Şu anda Türkçe odaklı - uluslararası kullanıcılar için gerekli"],
        ["Mobil Uygulama", "🟢 İsteğe Bağlı", "12", "Çok Yüksek", "Yüksek", "7/7", "iOS/Android - farklı platform, büyük yatırım gerektirir"],
    ]

    # Başlık satırını yaz
    ws.append(headers)

    # Veri satırlarını yaz
    for row in data:
        ws.append(row)

    # Stil tanımlamaları
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    critical_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    important_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    optional_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Başlık satırı formatla
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Veri satırlarını formatla
    for row_idx in range(2, len(data) + 2):
        priority = ws.cell(row=row_idx, column=2).value

        # Öncelik seviyesine göre satır rengi
        row_fill = None
        if "🔴 Kritik" in priority:
            row_fill = critical_fill
        elif "🟡 Önemli" in priority:
            row_fill = important_fill
        elif "🟢 İsteğe Bağlı" in priority:
            row_fill = optional_fill

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if row_fill:
                cell.fill = row_fill

            # İlk sütun (özellik adı) sola hizalı ve kalın
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)

            # Öncelik sütunu kalın ve büyük
            if col_idx == 2:
                cell.font = Font(bold=True, size=12)

            # Son sütun (açıklama) sola hizalı
            if col_idx == len(headers):
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Sütun genişliklerini ayarla
    column_widths = [30, 18, 15, 15, 18, 18, 55]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Satır yüksekliğini ayarla
    ws.row_dimensions[1].height = 30
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 35

    # Açıklama satırları ekle
    last_row = len(data) + 2
    ws.append([])
    ws.append(["Sembol Açıklaması:"])
    ws.cell(row=last_row + 2, column=1).font = Font(bold=True, size=11)

    ws.append(["🔴 Kritik = 1-2 ay içinde eklenirse rekabette kalır"])
    ws.append(["🟡 Önemli = 3-6 ay içinde eklenirse pazar payı artar"])
    ws.append(["🟢 İsteğe Bağlı = 6-12 ay içinde eklenirse farklılaşma sağlar"])
    ws.append([])
    ws.append(["Zorluk Seviyeleri:"])
    ws.cell(row=last_row + 7, column=1).font = Font(bold=True, size=11)

    ws.append(["Düşük = 1 geliştirici, basit entegrasyon"])
    ws.append(["Orta = 1 geliştirici, orta karmaşıklık"])
    ws.append(["Yüksek = 1 geliştirici veya dış kaynak, karmaşık"])
    ws.append(["Çok Yüksek = Ekip gerektirir veya uzmanlık alanı dışı"])
    ws.append([])
    ws.append(["Toplam Tahmini Süre (Kritik Özellikler): 10 hafta"])
    ws.append(["Toplam Tahmini Süre (Tüm Özellikler): 41 hafta"])

    # Kaydet
    wb.save("/mnt/c/Users/serkan/source/repos/QuadroAIPilot setup so so outlook not setup deneme2/QuadroAIPilot_Eksikler_Oncelik.xlsx")
    print("✅ QuadroAIPilot_Eksikler_Oncelik.xlsx oluşturuldu")

if __name__ == "__main__":
    print("📊 Excel raporları oluşturuluyor...")
    create_competitor_analysis()
    create_priority_analysis()
    print("\n✅ Tüm Excel dosyaları başarıyla oluşturuldu!")
    print("📁 Dosyalar: QuadroAIPilot_Rakip_Analizi.xlsx ve QuadroAIPilot_Eksikler_Oncelik.xlsx")
